import os
from urllib.parse import quote

from kivy.lang import Builder
from kivy.utils import platform
from kivy.uix.screenmanager import Screen, ScreenManager
from kivy.clock import Clock

from kivymd.app import MDApp
from kivymd.uix.button import MDRaisedButton
from kivymd.uix.dialog import MDDialog
from kivymd.uix.list import OneLineListItem

# Android intent support
if platform == "android":
    from jnius import autoclass

try:
    from openpyxl import load_workbook, Workbook
except Exception as e:
    load_workbook = None
    Workbook = None

KV = '''
ScreenManager:
    MainScreen:

<MainScreen>:
    name: "main"
    BoxLayout:
        orientation: "vertical"
        padding: 16
        spacing: 14

        GridLayout:
            cols: 2
            size_hint_y: None
            height: self.minimum_height
            spacing: 8
            MDTextField:
                id: entry_name
                hint_text: "Name"
                mode: "rectangle"
            MDTextField:
                id: entry_phone
                hint_text: "Phone"
                mode: "rectangle"
            MDTextField:
                id: entry_monthly
                hint_text: "Monthly"
                mode: "rectangle"
                input_filter: 'float'
            MDTextField:
                id: entry_payment
                hint_text: "Payment"
                mode: "rectangle"
                input_filter: 'float'
        MDRaisedButton:
            text: "Add Customer"
            size_hint_y: None
            height: "48dp"
            on_release: app.add_customer()
        MDLabel:
            text: "Customers List"
            bold: True
            font_style: "H6"
            halign: "center"
            size_hint_y: None
            height: "32dp"
        ScrollView:
            size_hint: (1, 0.45)
            MDList:
                id: customer_list
        BoxLayout:
            size_hint_y: None
            height: "56dp"
            spacing: 8
            MDTextField:
                id: entry_extra_payment
                hint_text: "Extra Payment"
                mode: "rectangle"
                input_filter: "float"
            MDRaisedButton:
                text: "Add Payment"
                on_release: app.add_payment()
        BoxLayout:
            size_hint_y: None
            height: "56dp"
            spacing: 8
            MDTextField:
                id: entry_extra_monthly
                hint_text: "Extra Monthly"
                mode: "rectangle"
                input_filter: "float"
            MDRaisedButton:
                text: "Increase Monthly"
                on_release: app.add_monthly()
        MDRaisedButton:
            text: "Remove Selected Customer"
            size_hint_y: None
            height: "48dp"
            md_bg_color: 1, 0.3, 0.3, 1
            on_release: app.confirm_remove_customer()
'''

class MainScreen(Screen):
    pass


def safe_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


class CustomerApp(MDApp):
    selected_index = None
    remove_dialog = None

    def build(self):
        self.theme_cls.primary_palette = "Blue"
        self.theme_cls.theme_style = "Light"
        return Builder.load_string(KV)

    @property
    def excel_file(self):
        # Use app-specific data dir on Android; desktop fallback works too
        return os.path.join(self.user_data_dir, "customers.xlsx")

    def on_start(self):
        # Ensure Excel file is present in a writable location
        self.ensure_excel_ready()
        self.customers = self.load_from_excel()
        Clock.schedule_once(lambda *_: self.refresh_list(), 0)

    def ensure_excel_ready(self):
        dest = self.excel_file
        if os.path.exists(dest):
            return
        # Try to copy from the script directory if provided alongside main.py
        src = os.path.join(os.path.dirname(os.path.abspath(__file__)), "customers.xlsx")
        if os.path.exists(src):
            try:
                import shutil
                shutil.copy2(src, dest)
                return
            except Exception:
                pass
        # Otherwise create a new workbook with headers
        if Workbook is None:
            self.show_dialog("openpyxl not available; cannot create Excel file.")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"
        ws.append(["Name", "Phone", "Monthly", "Payment", "Due"])
        wb.save(dest)

    def load_from_excel(self):
        path = self.excel_file
        if not os.path.isfile(path):
            return []
        if load_workbook is None:
            self.show_dialog("openpyxl not available; cannot read Excel file.")
            return []
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        # Read headers
        headers = [str(c.value) if c.value is not None else "" for c in ws[1]] if ws.max_row >= 1 else []
        idx = {h: i for i, h in enumerate(headers)}
        wanted = ["Name", "Phone", "Monthly", "Payment", "Due"]

        customers = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = (row[idx.get("Name", -1)] if idx.get("Name") is not None else "") if row else ""
            phone = (row[idx.get("Phone", -1)] if idx.get("Phone") is not None else "") if row else ""
            monthly = safe_float((row[idx.get("Monthly", -1)] if idx.get("Monthly") is not None else 0)) if row else 0.0
            payment = safe_float((row[idx.get("Payment", -1)] if idx.get("Payment") is not None else 0)) if row else 0.0
            due_raw = (row[idx.get("Due", -1)] if idx.get("Due") is not None else None) if row else None
            due = safe_float(due_raw) if due_raw is not None else (monthly - payment)
            customers.append({
                "Name": str(name or ""),
                "Phone": str(phone or ""),
                "Monthly": float(monthly),
                "Payment": float(payment),
                "Due": float(due),
            })
        # Ensure keys exist
        for c in customers:
            for k in wanted:
                if k not in c:
                    c[k] = "" if k in ("Name", "Phone") else 0.0
        return customers

    def save_to_excel(self):
        path = self.excel_file
        if Workbook is None:
            self.show_dialog("openpyxl not available; cannot write Excel file.")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"
        ws.append(["Name", "Phone", "Monthly", "Payment", "Due"])
        for c in self.customers:
            ws.append([c.get("Name", ""), c.get("Phone", ""), float(c.get("Monthly", 0)), float(c.get("Payment", 0)), float(c.get("Due", 0))])
        wb.save(path)

    def show_dialog(self, text):
        try:
            if hasattr(self, '_dialog') and self._dialog:
                self._dialog.dismiss()
            self._dialog = MDDialog(text=text, size_hint=(0.8, 0.2))
            self._dialog.open()
        except Exception:
            # Fallback: avoid crashing if dialog cannot be created yet
            pass

    def refresh_list(self):
        screen = self.root.get_screen("main")
        customer_list = screen.ids.customer_list
        customer_list.clear_widgets()
        self.list_items = []
        for i, c in enumerate(self.customers):
            line = f"Name: {c['Name']}, Phone: {c['Phone']}, Monthly: {c['Monthly']}, Payment: {c['Payment']}, Due: {c['Due']}"
            if i == self.selected_index:
                line += "  [selected]"
            item = OneLineListItem(text=line)
            item.bind(on_release=lambda widget, idx=i: self.select_customer(idx))
            self.list_items.append(item)
            customer_list.add_widget(item)

    def select_customer(self, index):
        self.selected_index = index
        self.refresh_list()

    def add_customer(self):
        ids = self.root.get_screen("main").ids
        name = ids.entry_name.text.strip()
        phone = ids.entry_phone.text.strip()
        monthly = ids.entry_monthly.text.strip()
        payment = ids.entry_payment.text.strip()
        if not name or not phone or not monthly or not payment:
            self.show_dialog("Please fill all fields.")
            return
        try:
            monthly = float(monthly)
            payment = float(payment)
        except ValueError:
            self.show_dialog("Monthly and Payment must be numbers.")
            return
        due = monthly - payment
        self.customers.append({
            "Name": name,
            "Phone": phone,
            "Monthly": monthly,
            "Payment": payment,
            "Due": due,
        })
        self.save_to_excel()
        ids.entry_name.text = ""
        ids.entry_phone.text = ""
        ids.entry_monthly.text = ""
        ids.entry_payment.text = ""
        self.selected_index = None
        self.refresh_list()

    def add_payment(self):
        ids = self.root.get_screen("main").ids
        if self.selected_index is None:
            self.show_dialog("Please select a customer.")
            return
        extra_pay = ids.entry_extra_payment.text.strip()
        if not extra_pay:
            self.show_dialog("Please enter payment amount.")
            return
        try:
            extra_pay = float(extra_pay)
            if extra_pay <= 0:
                raise ValueError
        except ValueError:
            self.show_dialog("Payment must be positive.")
            return
        c = self.customers[self.selected_index]
        c["Payment"] = float(c.get("Payment", 0)) + extra_pay
        c["Due"] = float(c.get("Monthly", 0)) - float(c["Payment"])
        self.save_to_excel()
        ids.entry_extra_payment.text = ""
        self.refresh_list()
        # WhatsApp message - send directly
        message = f"Thanks for your {extra_pay} payment and your remaining due is {c['Due']}"
        self.send_whatsapp_message(c.get("Phone", ""), message)

    def add_monthly(self):
        ids = self.root.get_screen("main").ids
        if self.selected_index is None:
            self.show_dialog("Please select a customer.")
            return
        extra_monthly = ids.entry_extra_monthly.text.strip()
        if not extra_monthly:
            self.show_dialog("Please enter monthly increment.")
            return
        try:
            extra_monthly = float(extra_monthly)
            if extra_monthly <= 0:
                raise ValueError
        except ValueError:
            self.show_dialog("Monthly increment must be positive.")
            return
        c = self.customers[self.selected_index]
        c["Monthly"] = float(c.get("Monthly", 0)) + extra_monthly
        c["Due"] = float(c["Monthly"]) - float(c.get("Payment", 0))
        self.save_to_excel()
        ids.entry_extra_monthly.text = ""
        self.refresh_list()

    def confirm_remove_customer(self):
        if self.selected_index is None or not self.customers:
            self.show_dialog("Please select a customer to remove.")
            return
        c = self.customers[self.selected_index]
        dialog_text = f"Are you sure you want to remove\nName: {c['Name']}, Phone: {c['Phone']}?"
        self.remove_dialog = MDDialog(
            text=dialog_text,
            buttons=[
                MDRaisedButton(
                    text="Cancel",
                    on_release=lambda _: self.remove_dialog.dismiss()
                ),
                MDRaisedButton(
                    text="Remove",
                    md_bg_color=(1, 0.3, 0.3, 1),
                    on_release=lambda _: self.remove_customer_confirmed()
                ),
            ],
        )
        self.remove_dialog.open()

    def remove_customer_confirmed(self):
        if self.remove_dialog:
            self.remove_dialog.dismiss()
            self.remove_dialog = None
        self.remove_customer()

    def remove_customer(self):
        if self.selected_index is not None and 0 <= self.selected_index < len(self.customers):
            self.customers.pop(self.selected_index)
            self.save_to_excel()
            self.selected_index = None
            self.refresh_list()

    def send_whatsapp_message(self, phone, message):
        p = str(phone or '').strip()
        p = ''.join(ch for ch in p if ch.isdigit() or ch == '+')
        if not p:
            self.show_dialog("No phone number provided.")
            return
        if p.startswith('0'):
            # Default country prefix (adjust to your locale)
            p = "+977" + p[1:]  # Nepal as in original
        elif not p.startswith('+'):
            self.show_dialog("Use international format (e.g., +977...).")
            return
        msg = quote(message)
        url = f"https://wa.me/{p}?text={msg}"
        if platform == "android":
            try:
                PythonActivity = autoclass('org.kivy.android.PythonActivity')
                activity = PythonActivity.mActivity
                Intent = autoclass('android.content.Intent')
                Uri = autoclass('android.net.Uri')
                intent = Intent(Intent.ACTION_VIEW)
                intent.setData(Uri.parse(url))
                activity.startActivity(intent)
            except Exception:
                # Fallback to webbrowser if intent fails
                import webbrowser
                webbrowser.open(url)
        else:
            import webbrowser
            webbrowser.open(url)


if __name__ == "__main__":
    CustomerApp().run()
